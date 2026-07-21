import openpyxl
import json
import sys
import os

def export_bom(excel_path, output_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    
    components = []
    item = 1
    for row in range(2, ws.max_row + 1):
        values = [cell.value for cell in ws[row]]
        if not values[0]:
            continue
        
        comp = {}
        for i, header in enumerate(headers):
            comp[header] = values[i]
        
        comp['item'] = item
        item += 1
        
        if 'visible3d' not in comp:
            comp['visible3d'] = True
        
        components.append(comp)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(components, f, indent= 4, ensure_ascii=False)

    print(f"Exported {len(components)} components to {output_path}")

if __name__ == '__main__':
    excel_path = sys.argv[1] if len(sys.argv) > 1 else 'resources/excel/ER5215-301.xlsx'
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'data/components.json'
    
    export_bom(excel_path, output_path)
